from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import unicodedata
import urllib.parse
import urllib.request
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = ROOT.parents[1]
DLC3_DIR = ROOT / "player_dlc3"
INPUT_LIST = DLC3_DIR / "list.txt"
IDENTITIES = DLC3_DIR / "dlc3-identities.json"
TEMPLATE = DLC3_DIR / "第二批DLC球员_S4耦合回归最终数据.xlsx"
OUTPUT_XLSX = DLC3_DIR / "DLC3球员_S4耦合回归初版.xlsx"
OUTPUT_JSON = DLC3_DIR / "DLC3球员_S4耦合回归初版.json"

S3_SOURCE = WORKSPACE_ROOT / "backups" / "player-data" / "S3-player-pool-700-20260725.json"
FIFA22_SOURCE = ROOT / "data" / "sources" / "kaggle-history-dlc2" / "fifa22" / "players_fifa22.csv"
FIFA21_SOURCE = ROOT / "data" / "sources" / "kaggle-history-dlc2" / "fifa11-21" / "fifa21_male2.csv"
FUT_DIR = ROOT / "data" / "sources" / "kaggle-history-dlc2" / "fut10-20"
CURRENT_POOL_SOURCE = ROOT / "data" / "s4-player-pool-with-dlc.json"
DLC1_SOURCE = ROOT / "data" / "player-dlc-s4-final.json"
DLC2_SOURCE = ROOT / "data" / "player-dlc2-s4-final.json"
CACHE_DIR = ROOT / "data" / "sources" / "ea-official-dlc3-20260808"

EA_API = "https://drop-api.ea.com/rating/ea-sports-fc"
EA_RATINGS = "https://www.ea.com/games/ea-sports-fc/ratings"
BUILD_DATE = "2026-08-08"
ATTRIBUTE_CAP = 96

ATTRS = [
    "passing", "firstTouch", "dribbling", "crossing", "finishing", "longShots", "heading", "setPieces",
    "tackling", "marking", "positioning", "vision", "decisions", "composure", "offBall", "discipline",
    "pace", "acceleration", "strength", "stamina", "agility", "jumping", "workRate", "aggression",
    "goalkeeping", "reflexes",
]

ATTR_LABELS = {
    "passing": "传球", "firstTouch": "停球", "dribbling": "盘带", "crossing": "传中",
    "finishing": "射门", "longShots": "远射", "heading": "头球", "setPieces": "定位球",
    "tackling": "抢断", "marking": "盯人", "positioning": "站位", "vision": "视野",
    "decisions": "决策", "composure": "冷静", "offBall": "无球", "discipline": "纪律",
    "pace": "速度", "acceleration": "加速", "strength": "力量", "stamina": "耐力",
    "agility": "灵活", "jumping": "弹跳", "workRate": "投入", "aggression": "侵略性",
    "goalkeeping": "守门", "reflexes": "反应",
}

OVERALL_KEYS = {
    "GK": ["goalkeeping", "reflexes", "positioning", "composure"],
    "DEF": ["tackling", "marking", "positioning", "strength", "pace"],
    "MID": ["passing", "vision", "decisions", "firstTouch", "stamina"],
    "ATT": ["finishing", "offBall", "pace", "dribbling", "composure"],
}

FIFA22_OVERRIDES = {
    "Franck Ribéry": "156616", "Lukas Podolski": "150516", "Zlatan Ibrahimović": "41236",
    "Leonardo Bonucci": "184344", "David Silva": "168542", "Jesús Navas": "146536",
    "Juan Mata": "178088",
}

FIFA21_OVERRIDES = {
    "Adriano": "106019", "Diego Costa": "179844", "Gilberto Silva": "47390",
    "Víctor Valdés": "106573", "David Silva": "168542", "Jesús Navas": "146536",
    "Juan Mata": "178088",
}


def norm_name(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]", "", ascii_name)


def norm_chinese(value):
    return re.sub(r"[·•・\s（）()\-]", "", str(value or ""))


def number(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def rounded(value, minimum=1, maximum=ATTRIBUTE_CAP):
    return max(minimum, min(maximum, int(round(value))))


def avg(*values):
    valid = [value for value in values if isinstance(value, (int, float)) and math.isfinite(value)]
    return sum(valid) / len(valid) if valid else None


def role_pool(role):
    if role == "GK":
        return "GK"
    if role in {"CB", "LB", "RB"}:
        return "DEF"
    if role in {"DM", "AM", "LM", "RM"}:
        return "MID"
    return "ATT"


def map_role(value):
    return {
        "GK": "GK", "CB": "CB", "LB": "LB", "LWB": "LB", "RB": "RB", "RWB": "RB",
        "CDM": "DM", "CM": "AM", "CAM": "AM", "LM": "LM", "RM": "RM",
        "ST": "ST", "CF": "ST", "LF": "LW", "RF": "RW", "LW": "LW", "RW": "RW",
    }.get(str(value or "").strip(), "")


def grade_for(overall):
    if overall >= 90:
        return "S"
    if overall >= 86:
        return "A"
    if overall >= 80:
        return "B"
    return "C"


def load_json(path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def player_array(payload):
    if isinstance(payload, list):
        return payload
    return payload.get("players", [])


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_positions(value):
    return [entry.strip() for entry in re.sub(r"[\[\]'\"]", "", str(value or "")).split(",") if entry.strip()]


def parse_feet_height(value):
    match = re.search(r"(\d+)'(\d+)", str(value or ""))
    return round((int(match.group(1)) * 12 + int(match.group(2))) * 2.54) if match else None


def current_attributes(player):
    attributes = player.get("attributes") or {}
    return {key: number(attributes.get(key)) for key in ATTRS}


def fifa_raw_to_core(source, role):
    goalkeeper = role == "GK"
    defensive = role in {"CB", "LB", "RB", "DM"}
    raw = {
        "passing": avg(source.get("passing"), source.get("shortPassing"), source.get("longPassing")),
        "firstTouch": source.get("ballControl"), "dribbling": source.get("dribbling"),
        "crossing": source.get("crossing"), "finishing": source.get("finishing"),
        "longShots": source.get("longShots"), "heading": source.get("headingAccuracy"),
        "setPieces": avg(source.get("freeKickAccuracy"), source.get("curve"), source.get("penalties")),
        "tackling": avg(source.get("standingTackle"), source.get("slidingTackle")),
        "marking": source.get("defensiveAwareness"),
        "positioning": source.get("goalkeeperPositioning") if goalkeeper else source.get("defensiveAwareness") if defensive else avg(source.get("defensiveAwareness"), source.get("interceptions")),
        "vision": source.get("vision"), "decisions": source.get("reactions"),
        "composure": source.get("composure"), "offBall": source.get("positioning"),
        "discipline": avg(source.get("composure"), source.get("reactions"), 105 - source["aggression"] if number(source.get("aggression")) is not None else None),
        "pace": source.get("pace"), "acceleration": source.get("acceleration"),
        "strength": source.get("strength"), "stamina": source.get("stamina"),
        "agility": source.get("agility"), "jumping": source.get("jumping"),
        "workRate": source.get("stamina"), "aggression": source.get("aggression"),
        "goalkeeping": avg(source.get("goalkeeperDiving"), source.get("goalkeeperHandling"), source.get("goalkeeperKicking"), source.get("goalkeeperPositioning"), source.get("goalkeeperReflexes")) if goalkeeper else 8,
        "reflexes": source.get("goalkeeperReflexes") if goalkeeper else 8,
    }
    return {key: rounded(value, 35 if key == "discipline" else 1, 95 if key == "discipline" else 99) if number(value) is not None else None for key, value in raw.items()}


def from_s3(player):
    return {
        "version": "YellowDogs S3 EA-derived archive",
        "type": "project S3 EA-derived historical profile",
        "overall": number(player.get("overall")), "sourceId": player.get("id"),
        "sourceUrl": "", "datasetUrl": "backups/player-data/S3-player-pool-700-20260725.json",
        "attributes": current_attributes(player), "inputCoverage": 26,
    }


def from_current(player, source_label):
    return {
        "version": source_label, "type": "current S4 player pool",
        "overall": number(player.get("overall") or player.get("proposedOverall")),
        "sourceId": player.get("id"), "sourceUrl": player.get("sourceUrl") or "",
        "datasetUrl": player.get("source") or "data/s4-player-pool-with-dlc.json",
        "attributes": current_attributes(player), "inputCoverage": 26,
    }


def raw_fifa22(row):
    return {
        "pace": number(row.get("PaceTotal")), "acceleration": number(row.get("Acceleration")),
        "positioning": number(row.get("Positioning")), "finishing": number(row.get("Finishing")),
        "longShots": number(row.get("LongShots")), "vision": number(row.get("Vision")),
        "crossing": number(row.get("Crossing")), "freeKickAccuracy": number(row.get("FKAccuracy")),
        "shortPassing": number(row.get("ShortPassing")), "longPassing": number(row.get("LongPassing")),
        "curve": number(row.get("Curve")), "dribbling": number(row.get("Dribbling")),
        "agility": number(row.get("Agility")), "reactions": number(row.get("Reactions")),
        "ballControl": number(row.get("BallControl")), "composure": number(row.get("Composure")),
        "interceptions": number(row.get("Interceptions")), "headingAccuracy": number(row.get("HeadingAccuracy")),
        "defensiveAwareness": number(row.get("Marking")), "standingTackle": number(row.get("StandingTackle")),
        "slidingTackle": number(row.get("SlidingTackle")), "jumping": number(row.get("Jumping")),
        "stamina": number(row.get("Stamina")), "strength": number(row.get("Strength")),
        "aggression": number(row.get("Aggression")), "penalties": number(row.get("Penalties")),
        "goalkeeperDiving": number(row.get("GKDiving")), "goalkeeperHandling": number(row.get("GKHandling")),
        "goalkeeperKicking": number(row.get("GKKicking")), "goalkeeperPositioning": number(row.get("GKPositioning")),
        "goalkeeperReflexes": number(row.get("GKReflexes")), "passing": number(row.get("PassingTotal")),
    }


def raw_fifa21(row):
    return {
        "pace": number(row.get("PAC")), "acceleration": number(row.get("Acceleration")),
        "positioning": number(row.get("Positioning")), "finishing": number(row.get("Finishing")),
        "longShots": number(row.get("Long Shots")), "vision": number(row.get("Vision")),
        "crossing": number(row.get("Crossing")), "freeKickAccuracy": number(row.get("FK Accuracy")),
        "shortPassing": number(row.get("Short Passing")), "longPassing": number(row.get("Long Passing")),
        "curve": number(row.get("Curve")), "dribbling": number(row.get("Dribbling")),
        "agility": number(row.get("Agility")), "reactions": number(row.get("Reactions")),
        "ballControl": number(row.get("Ball Control")), "composure": number(row.get("Composure")),
        "interceptions": number(row.get("Interceptions")), "headingAccuracy": number(row.get("Heading Accuracy")),
        "defensiveAwareness": number(row.get("Marking")), "standingTackle": number(row.get("Standing Tackle")),
        "slidingTackle": number(row.get("Sliding Tackle")), "jumping": number(row.get("Jumping")),
        "stamina": number(row.get("Stamina")), "strength": number(row.get("Strength")),
        "aggression": number(row.get("Aggression")), "penalties": number(row.get("Penalties")),
        "goalkeeperDiving": number(row.get("GK Diving")), "goalkeeperHandling": number(row.get("GK Handling")),
        "goalkeeperKicking": number(row.get("GK Kicking")), "goalkeeperPositioning": number(row.get("GK Positioning")),
        "goalkeeperReflexes": number(row.get("GK Reflexes")), "passing": number(row.get("PAS")),
    }


def from_fifa22(row, role):
    return {
        "version": "FIFA 22 historical snapshot", "type": "Kaggle FIFA 22 complete player archive",
        "overall": number(row.get("Overall")), "sourceId": row.get("ID"), "sourceUrl": row.get("PhotoUrl") or "",
        "datasetUrl": "https://www.kaggle.com/datasets/cashncarry/fifa-22-complete-player-dataset",
        "attributes": fifa_raw_to_core(raw_fifa22(row), role), "inputCoverage": 26,
    }


def from_fifa21(row, role):
    season = re.search(r"\d{4}", str(row.get("Contract") or ""))
    return {
        "version": f"FIFA historical snapshot ({season.group(0) if season else 'legacy'})",
        "type": "Kaggle FIFA historical complete player archive", "overall": number(row.get("OVA")),
        "sourceId": row.get("ID"), "sourceUrl": row.get("Player Photo") or "",
        "datasetUrl": "https://www.kaggle.com/datasets/ekrembayar/fifa-21-complete-player-dataset",
        "attributes": fifa_raw_to_core(raw_fifa21(row), role), "inputCoverage": sum(number(value) is not None for value in fifa_raw_to_core(raw_fifa21(row), role).values()),
    }


def from_fut(row, file_name):
    values = {key: number(row.get(key)) for key in ("PAC", "SHO", "PAS", "DRI", "DEF", "PHY")}
    return {
        "version": f"{file_name.replace(' Fut Players.csv', '')} FUT historical card",
        "type": "Kaggle FUT historical six-stat card anchor", "overall": number(row.get("Ratings")),
        "sourceId": "", "sourceUrl": "",
        "datasetUrl": "https://www.kaggle.com/datasets/mohammedessam97/fifa-1020-fut-players-dataset",
        "attributes": {}, "fut": values, "inputCoverage": sum(value is not None for value in values.values()),
    }


def official_core(item, fallback_role):
    role = map_role((item.get("position") or {}).get("shortLabel")) or fallback_role
    stats = item.get("stats") or {}
    value = lambda key: number((stats.get(key) or {}).get("value"))
    raw = {
        "pace": value("pac"), "acceleration": value("acceleration"), "positioning": value("positioning"),
        "finishing": value("finishing"), "longShots": value("longShots"), "vision": value("vision"),
        "crossing": value("crossing"), "freeKickAccuracy": value("freeKickAccuracy"),
        "shortPassing": value("shortPassing"), "longPassing": value("longPassing"), "curve": value("curve"),
        "dribbling": value("dribbling"), "agility": value("agility"), "reactions": value("reactions"),
        "ballControl": value("ballControl"), "composure": value("composure"), "interceptions": value("interceptions"),
        "headingAccuracy": value("headingAccuracy"), "defensiveAwareness": value("defensiveAwareness"),
        "standingTackle": value("standingTackle"), "slidingTackle": value("slidingTackle"),
        "jumping": value("jumping"), "stamina": value("stamina"), "strength": value("strength"),
        "aggression": value("aggression"), "penalties": value("penalties"),
        "goalkeeperDiving": value("gkDiving"), "goalkeeperHandling": value("gkHandling"),
        "goalkeeperKicking": value("gkKicking"), "goalkeeperPositioning": value("gkPositioning"),
        "goalkeeperReflexes": value("gkReflexes"), "passing": value("pas"),
    }
    return fifa_raw_to_core(raw, role)


def official_name(item):
    return item.get("commonName") or f"{item.get('firstName', '')} {item.get('lastName', '')}".strip()


def fetch_official(identity, order):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache = CACHE_DIR / f"{order:02d}-{norm_name(identity['nameEn'])}.json"
    if cache.exists():
        payload = load_json(cache)
    else:
        query = urllib.parse.urlencode({"locale": "en", "limit": 20, "search": identity["nameEn"]})
        request = urllib.request.Request(f"{EA_API}?{query}", headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(request, timeout=30) as response:
            payload = json.loads(response.read().decode("utf-8"))
        cache.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    items = payload.get("items") or []
    expected = norm_name(identity["nameEn"])
    exact = next((item for item in items if any(norm_name(value) == expected for value in [item.get("commonName"), f"{item.get('firstName', '')} {item.get('lastName', '')}".strip()])), None)
    return exact, [official_name(item) for item in items[:5]], int(payload.get("totalItems") or 0)


def training_samples(template_workbook):
    samples = []
    seen = set()
    for source in [CURRENT_POOL_SOURCE, DLC1_SOURCE, DLC2_SOURCE]:
        if not source.exists():
            continue
        for player in player_array(load_json(source)):
            attributes = current_attributes(player)
            role = player.get("role") or ""
            overall = number(player.get("overall") or player.get("proposedOverall"))
            key = player.get("id") or f"{player.get('sourceName')}:{role}:{overall}"
            if key in seen or role not in {"GK", "CB", "LB", "RB", "DM", "AM", "LM", "RM", "ST", "LW", "RW"} or overall is None or any(value is None for value in attributes.values()):
                continue
            seen.add(key)
            samples.append({"name": player.get("name") or player.get("displayNameZh") or player.get("sourceName") or key, "role": role, "secondaryRole": player.get("secondaryRole") or "", "overall": overall, "heightCm": number(player.get("heightCm")), "attributes": attributes})
    sheet = template_workbook["球员数据"]
    for row in range(2, sheet.max_row + 1):
        role = sheet.cell(row, 18).value
        overall = number(sheet.cell(row, 15).value)
        values = {key: number(sheet.cell(row, 25 + index).value) for index, key in enumerate(ATTRS)}
        if role and overall is not None and all(value is not None for value in values.values()):
            samples.append({"name": sheet.cell(row, 4).value, "role": role, "secondaryRole": sheet.cell(row, 19).value or "", "overall": overall, "heightCm": number(sheet.cell(row, 22).value), "attributes": values})
    return samples


def regression(identity, samples):
    role_samples = [sample for sample in samples if sample["role"] == identity["role"]]
    def distance(sample):
        value = abs(sample["overall"] - identity["targetOverall"])
        if sample["heightCm"] is not None:
            value += abs(sample["heightCm"] - identity["heightCm"]) * 0.08
        if identity.get("secondaryRole") and sample.get("secondaryRole") == identity["secondaryRole"]:
            value -= 0.35
        return max(0.15, value)
    neighbors = sorted(role_samples, key=distance)[:5]
    predicted = {}
    for key in ATTRS:
        weighted = []
        for sample in neighbors:
            adjusted = sample["attributes"][key] + (identity["targetOverall"] - sample["overall"]) * 0.65
            weighted.append((adjusted, 1 / (0.5 + distance(sample))))
        predicted[key] = rounded(sum(value * weight for value, weight in weighted) / sum(weight for _, weight in weighted))
    return predicted, [sample["name"] for sample in neighbors]


def blend_source(predicted, source, identity):
    original = source.get("attributes") or {}
    final = dict(predicted)
    source_overall = source.get("overall")
    delta = (identity["targetOverall"] - source_overall) * 0.65 if source_overall is not None else 0
    for key in ATTRS:
        value = number(original.get(key))
        if value is not None:
            final[key] = rounded((value + delta) * 0.82 + predicted[key] * 0.18)
    fut = source.get("fut") or {}
    anchors = {
        "PAC": ["pace", "acceleration"], "SHO": ["finishing", "longShots"],
        "PAS": ["passing", "vision"], "DRI": ["dribbling", "firstTouch"],
        "DEF": ["tackling", "marking"], "PHY": ["strength", "aggression"],
    }
    for stat, keys in anchors.items():
        value = number(fut.get(stat))
        if value is None:
            continue
        adjusted = value + delta
        for key in keys:
            final[key] = rounded(adjusted * 0.55 + predicted[key] * 0.45)
    return final


def couple_overall(attributes, role, target):
    keys = OVERALL_KEYS[role_pool(role)]
    required = int(target * len(keys))
    current = sum(int(attributes[key]) for key in keys)
    delta = required - current
    priority = list(keys)
    cursor = 0
    while delta and cursor < 500:
        key = priority[cursor % len(priority)]
        direction = 1 if delta > 0 else -1
        proposed = attributes[key] + direction
        if 1 <= proposed <= ATTRIBUTE_CAP:
            attributes[key] = proposed
            delta -= direction
        cursor += 1
    if delta:
        raise RuntimeError(f"cannot couple {role} to {target}")
    return attributes


def calculated_overall(attributes, role):
    keys = OVERALL_KEYS[role_pool(role)]
    return round(sum(attributes[key] for key in keys) / len(keys))


def wiki_search(name):
    return "https://en.wikipedia.org/w/index.php?search=" + urllib.parse.quote(name)


def copy_row_style(sheet, source_row, target_row, max_column):
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def replace_sheet_rows(sheet, rows, template_row=2):
    styles = [copy(sheet.cell(template_row, column)._style) for column in range(1, sheet.max_column + 1)]
    alignments = [copy(sheet.cell(template_row, column).alignment) for column in range(1, sheet.max_column + 1)]
    row_height = sheet.row_dimensions[template_row].height
    if sheet.max_row >= 2:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            if column <= len(styles):
                cell._style = copy(styles[column - 1])
                cell.alignment = copy(alignments[column - 1])
        sheet.row_dimensions[row_index].height = row_height


def main():
    requested_names = [line.strip() for line in INPUT_LIST.read_text(encoding="utf-8-sig").splitlines() if line.strip()]
    identities = load_json(IDENTITIES)
    if requested_names != [entry["input"] for entry in identities]:
        raise RuntimeError("list.txt 与 dlc3-identities.json 的名单或顺序不一致")

    workbook = load_workbook(TEMPLATE)
    samples = training_samples(workbook)
    s3_players = player_array(load_json(S3_SOURCE))
    current_players = player_array(load_json(CURRENT_POOL_SOURCE))
    dlc2_players = player_array(load_json(DLC2_SOURCE))
    duplicate_players = current_players + dlc2_players
    fifa22_rows = read_csv(FIFA22_SOURCE)
    fifa21_rows = read_csv(FIFA21_SOURCE)
    fifa22_by_id = {str(row.get("ID")): row for row in fifa22_rows}
    fifa21_by_id = {str(row.get("ID")): row for row in fifa21_rows}
    fifa22_by_name = {norm_name(row.get("FullName")): row for row in fifa22_rows}
    fifa21_by_name = {norm_name(row.get("Name")): row for row in fifa21_rows}

    fut_index = {}
    for path in FUT_DIR.glob("*.csv"):
        for row in read_csv(path):
            fut_index.setdefault(norm_name(row.get("Name")), []).append((row, path.name))

    results = []
    for index, identity in enumerate(identities, start=1):
        exact, candidates, official_total = fetch_official(identity, index)
        duplicate = next((player for player in duplicate_players if norm_name(player.get("sourceName")) == norm_name(identity["nameEn"])), None)
        s3 = next((player for player in s3_players if norm_chinese(player.get("name")) in {norm_chinese(identity["input"]), norm_chinese(identity["nameZh"])}), None)
        fifa22 = fifa22_by_id.get(FIFA22_OVERRIDES.get(identity["nameEn"], "")) or fifa22_by_name.get(norm_name(identity["nameEn"]))
        fifa21 = fifa21_by_id.get(FIFA21_OVERRIDES.get(identity["nameEn"], "")) or fifa21_by_name.get(norm_name(identity["nameEn"]))
        fut_matches = fut_index.get(norm_name(identity["nameEn"]), [])
        fut = min(fut_matches, key=lambda entry: abs((number(entry[0].get("Ratings")) or 0) - identity["targetOverall"])) if fut_matches else None

        if exact:
            source = {
                "version": "EA SPORTS FC official live ratings", "type": "EA official Drop API",
                "overall": number(exact.get("overallRating")), "sourceId": exact.get("id"),
                "sourceUrl": f"{EA_RATINGS}/player-ratings/{urllib.parse.quote(official_name(exact).lower().replace(' ', '-'))}/{exact.get('id')}",
                "datasetUrl": EA_RATINGS, "attributes": official_core(exact, identity["role"]), "inputCoverage": 26,
            }
        elif duplicate:
            source = from_current(duplicate, "Current YellowDogs League S4")
        elif s3:
            source = from_s3(s3)
        elif fifa22:
            source = from_fifa22(fifa22, identity["role"])
        elif fifa21:
            source = from_fifa21(fifa21, identity["role"])
        elif fut:
            source = from_fut(*fut)
        else:
            source = {"version": f"YellowDogs S4 positional regression {BUILD_DATE}", "type": "S4 same-position regression model", "overall": None, "sourceId": "", "sourceUrl": "", "datasetUrl": "", "attributes": {}, "inputCoverage": 0}

        predicted, neighbors = regression(identity, samples)
        final_attributes = couple_overall(blend_source(predicted, source, identity), identity["role"], identity["targetOverall"])
        input_attributes = [number(value) for value in (source.get("attributes") or {}).values() if number(value) is not None]
        known_deltas = []
        for key, value in (source.get("attributes") or {}).items():
            numeric = number(value)
            if numeric is not None and key in final_attributes:
                known_deltas.append(abs(final_attributes[key] - numeric))

        identity_status = identity.get("identityReview") or "已确认"
        duplicate_status = "当前球员池已存在" if duplicate else "可新增候选"
        processing = "来源耦合回归" if source["inputCoverage"] >= 26 else "来源锚点+回归补齐" if source["inputCoverage"] else "同位置回归补齐"
        official_match = "精确命中" if exact else "仅相似结果" if official_total else "无结果"
        source_note = "EA 官方 FC Ratings 实时接口精确命中" if exact else "EA 当前接口未收录；使用当前 S4 已有球员档案" if duplicate else "EA 当前接口未收录；使用项目 S3 的 EA 派生历史档案" if s3 else "EA 当前接口未收录；使用 FIFA 22 完整历史快照" if fifa22 else "EA 当前接口未收录；使用 FIFA 历史完整数据库" if fifa21 else "EA 当前接口未收录；使用 FUT 历史六维卡面作为回归锚点" if fut else "EA 当前接口及本地历史档案均未找到可靠细项，使用同位置回归生成初值"
        note = (
            f"{source_note}；巅峰赛季：{identity['peakSeason']}，年龄{identity['peakAge']}岁，效力{identity['peakClub']}，"
            f"惯用{'左脚' if identity['preferredFoot'] == 'left' else '右脚'}；{processing}；目标OVR {identity['targetOverall']}；"
            f"输入覆盖{source['inputCoverage']}；近邻：{'、'.join(neighbors)}"
        )
        if identity.get("identityReview"):
            note = f"{identity['identityReview']}；{note}"
        if duplicate:
            note = f"重复提醒：当前池已存在 {duplicate.get('name') or duplicate.get('displayNameZh') or duplicate.get('sourceName')}（{duplicate.get('id') or 'DLC2记录'}）；{note}"

        results.append({
            **identity, "order": index, "identityStatus": identity_status,
            "officialMatch": official_match, "officialCandidates": " | ".join(candidates),
            "duplicateStatus": duplicate_status, "existingPlayerId": duplicate.get("id", "") if duplicate else "",
            "existingPlayerName": (duplicate.get("name") or duplicate.get("displayNameZh") or duplicate.get("sourceName") or "") if duplicate else "",
            "source": source, "attributes": final_attributes, "grade": grade_for(identity["targetOverall"]),
            "pool": role_pool(identity["role"]), "processing": processing,
            "inputCoverage": source["inputCoverage"], "regressionFields": 26 - min(26, source["inputCoverage"]),
            "neighbors": neighbors, "calculatedOverall": calculated_overall(final_attributes, identity["role"]),
            "knownAverageDelta": round(sum(known_deltas) / len(known_deltas), 3) if known_deltas else None,
            "capCount": sum(value == ATTRIBUTE_CAP for value in final_attributes.values()), "note": note,
        })

    if any(result["calculatedOverall"] != result["targetOverall"] for result in results):
        raise RuntimeError("存在属性计算 OVR 与目标不一致")

    build_workbook(workbook, results, samples)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    OUTPUT_JSON.write_text(json.dumps({"schemaVersion": 1, "generatedAt": BUILD_DATE, "officialApi": EA_API, "players": results}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "xlsx": str(OUTPUT_XLSX), "json": str(OUTPUT_JSON), "players": len(results),
        "officialExact": sum(result["officialMatch"] == "精确命中" for result in results),
        "duplicates": [result["input"] for result in results if result["duplicateStatus"] != "可新增候选"],
        "fullSource": sum(result["inputCoverage"] >= 26 for result in results),
        "partialSource": sum(0 < result["inputCoverage"] < 26 for result in results),
        "regressionOnly": sum(result["inputCoverage"] == 0 for result in results),
        "identityReview": [result["input"] for result in results if result["identityStatus"] != "已确认"],
    }, ensure_ascii=False, indent=2))


def build_workbook(workbook, results, samples):
    summary = workbook["说明与汇总"]
    summary["A1"] = "第三批 DLC 球员：S4 耦合回归初版"
    summary["A3"] = "初版名单人数"
    summary["B3"] = len(results)
    summary["A4"] = "重复候选"
    summary["B4"] = "、".join(result["input"] for result in results if result["duplicateStatus"] != "可新增候选") or "无"
    summary["D4"] = "初版以名单身份映射和目标 OVR 为待审核口径。EA 官方 FC Ratings 接口优先；当前接口不收录 Icons、Heroes 与大部分退役球员，因此依次使用当前 S4、项目 S3 EA 派生档案、FIFA 历史完整库、FUT 历史六维卡及同位置局部回归。最后按当前 S4 引擎核心属性做有界耦合，确保属性计算 OVR 与目标一致。基础属性上限为 96。"
    summary["A5"], summary["B5"] = "纯同位置回归", sum(result["inputCoverage"] == 0 for result in results)
    summary["A6"], summary["B6"] = "来源+回归混合", sum(0 < result["inputCoverage"] < 26 for result in results)
    summary["A7"], summary["B7"] = "完整来源耦合", sum(result["inputCoverage"] >= 26 for result in results)
    grades = [sum(result["grade"] == grade for result in results) for grade in ["S", "A", "B", "C"]]
    summary["A8"], summary["B8"] = "S / A / B / C", " / ".join(map(str, grades))
    summary["A9"], summary["B9"] = "属性上限", ATTRIBUTE_CAP
    summary["A10"], summary["B10"] = "生成日期", BUILD_DATE
    summary["A11"] = "数据边界"
    summary["B12"] = EA_API
    summary["B13"] = "FIFA 22 完整库、FIFA 历史完整库、FIFA 10–20 FUT 历史卡；逐行来源保留在球员数据表"
    summary["B14"] = f"当前 S4、前两批 DLC 与模板样本，共 {len(samples)} 个可用训练样本；按主位置、OVR、身高和副位置选择近邻"
    summary["B15"] = "这是供人工审核的初版；缺失值由同位置回归生成，不代表外部历史事实；儒尼奥尔、德容身份与全部目标 OVR 仍需审核"

    player_headers = [
        "序号", "分组", "名单名称", "中文全名", "英文名", "身份确认", "数据状态", "EA官方匹配", "官方候选",
        "重复状态", "现有球员ID", "现有球员名", "数据版本", "来源类型", "调整后OVR", "评级", "位置池", "主位置", "副位置", "国籍", "俱乐部", "身高cm", "年龄", "惯用脚",
        *[ATTR_LABELS[key] for key in ATTRS], "来源ID", "来源URL", "数据集URL", "备注",
    ]
    player_rows = []
    for result in results:
        player_rows.append([
            result["order"], result["nationality"], result["input"], result["nameZh"], result["nameEn"], result["identityStatus"],
            "26项已耦合回归", result["officialMatch"], result["officialCandidates"] or None, result["duplicateStatus"],
            result["existingPlayerId"] or None, result["existingPlayerName"] or None, result["source"]["version"], result["source"]["type"],
            result["targetOverall"], result["grade"], result["pool"], result["role"], result.get("secondaryRole") or None,
            result["nationality"], result["peakClub"], result["heightCm"], result["peakAge"], "左脚" if result["preferredFoot"] == "left" else "右脚",
            *[result["attributes"][key] for key in ATTRS], result["source"].get("sourceId") or None,
            result["source"].get("sourceUrl") or None, result["source"].get("datasetUrl") or None, result["note"],
        ])
    players_sheet = workbook["球员数据"]
    for column, value in enumerate(player_headers, start=1):
        players_sheet.cell(1, column, value)
    replace_sheet_rows(players_sheet, player_rows)
    players_sheet.data_validations.dataValidation = []
    last_row = len(results) + 1
    validations = [
        ("list", '"S,A,B,C,"', f"P2:P{last_row}"), ("list", '"GK,DEF,MID,ATT,"', f"Q2:Q{last_row}"),
        ("list", '"GK,CB,LB,RB,DM,AM,LM,RM,ST,LW,RW,"', f"R2:S{last_row}"), ("list", '"左脚,右脚"', f"X2:X{last_row}"),
    ]
    for validation_type, formula, sqref in validations:
        validation = DataValidation(type=validation_type, formula1=formula, allow_blank=True)
        players_sheet.add_data_validation(validation)
        validation.add(sqref)
    number_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="96", allow_blank=False)
    players_sheet.add_data_validation(number_validation)
    number_validation.add(f"O2:O{last_row}")
    number_validation.add(f"Y2:AX{last_row}")

    source_rows = [[
        result["order"], result["input"], result["nationality"], result["processing"], result["inputCoverage"],
        result["regressionFields"], result["source"].get("sourceId") or None, "、".join(result["neighbors"]), result["note"],
    ] for result in results]
    replace_sheet_rows(workbook["来源与异常"], source_rows)

    peak_rows = [[
        result["nameZh"], result["targetOverall"], result["peakSeason"], result["peakAge"], result["peakClub"],
        "左脚" if result["preferredFoot"] == "left" else "右脚", result["source"].get("sourceUrl") or wiki_search(result["nameEn"]),
    ] for result in results]
    replace_sheet_rows(workbook["巅峰口径"], peak_rows)

    audit_rows = [[
        result["order"], result["input"], result["role"], result.get("secondaryRole") or None, result["targetOverall"],
        result["calculatedOverall"], result["inputCoverage"], result["regressionFields"], result["source"].get("overall"),
        result["processing"], result["knownAverageDelta"], result["capCount"], "、".join(result["neighbors"]),
    ] for result in results]
    audit_sheet = workbook["回归审计"]
    replace_sheet_rows(audit_sheet, audit_rows)
    start = len(results) + 4
    audit_sheet.cell(start, 1, "位置")
    audit_sheet.cell(start, 2, "训练样本数")
    audit_sheet.cell(start, 3, "OVR最低")
    audit_sheet.cell(start, 4, "OVR最高")
    for offset, role in enumerate(["ST", "CB", "DM", "RW", "GK", "AM", "LB", "RB", "LW", "LM"], start=1):
        role_samples = [sample for sample in samples if sample["role"] == role]
        audit_sheet.cell(start + offset, 1, role)
        audit_sheet.cell(start + offset, 2, len(role_samples))
        audit_sheet.cell(start + offset, 3, min((sample["overall"] for sample in role_samples), default=None))
        audit_sheet.cell(start + offset, 4, max((sample["overall"] for sample in role_samples), default=None))


if __name__ == "__main__":
    main()
