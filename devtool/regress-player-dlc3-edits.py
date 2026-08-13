from __future__ import annotations

import importlib.util
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DLC3_DIR = ROOT / "player_dlc3"
INPUT_XLSX = DLC3_DIR / "DLC3球员_S4耦合回归初版.xlsx"
BASE_JSON = DLC3_DIR / "DLC3球员_S4耦合回归初版.json"
OUTPUT_XLSX = DLC3_DIR / "DLC3球员_S4耦合回归复核版.xlsx"
OUTPUT_JSON = DLC3_DIR / "DLC3球员_S4耦合回归复核版.json"
HELPER_PATH = ROOT / "devtool" / "build-player-dlc3-initial.py"


spec = importlib.util.spec_from_file_location("dlc3_builder", HELPER_PATH)
builder = importlib.util.module_from_spec(spec)
spec.loader.exec_module(builder)


def main():
    workbook = load_workbook(INPUT_XLSX)
    base_payload = builder.load_json(BASE_JSON)
    base_by_name = {player["input"]: player for player in base_payload["players"]}
    samples = builder.training_samples(load_workbook(builder.TEMPLATE))
    player_sheet = workbook.worksheets[1]
    current_rows = []

    for row_number in range(2, player_sheet.max_row + 1):
        input_name = player_sheet.cell(row_number, 3).value
        if not input_name:
            continue
        if input_name not in base_by_name:
            raise RuntimeError(f"找不到初版来源记录：{input_name}")
        values = [player_sheet.cell(row_number, column).value for column in range(1, 55)]
        current_rows.append((row_number, values))

    results = []
    player_rows = []
    for order, (_, values) in enumerate(current_rows, start=1):
        input_name = values[2]
        base = base_by_name[input_name]
        target_overall = int(values[14])
        role = values[17]
        secondary_role = values[18] or ""
        preferred_foot = "left" if values[23] == "左脚" else "right"
        identity = {
            "input": input_name,
            "nameZh": values[3],
            "nameEn": values[4],
            "nationality": values[19],
            "role": role,
            "secondaryRole": secondary_role,
            "heightCm": int(values[21]),
            "preferredFoot": preferred_foot,
            "targetOverall": target_overall,
            "peakSeason": base["peakSeason"],
            "peakAge": int(values[22]),
            "peakClub": values[20],
        }
        source = base["source"]
        predicted, neighbors = builder.regression(identity, samples)
        attributes = builder.couple_overall(
            builder.blend_source(predicted, source, identity),
            role,
            target_overall,
        )
        input_coverage = int(base.get("inputCoverage", source.get("inputCoverage", 0)))
        processing = "来源耦合回归" if input_coverage >= 26 else "来源锚点+回归补齐" if input_coverage else "同位置回归补齐"
        known_deltas = []
        for key, source_value in (source.get("attributes") or {}).items():
            numeric = builder.number(source_value)
            if numeric is not None and key in attributes:
                known_deltas.append(abs(attributes[key] - numeric))
        note = re.sub(r"目标OVR\s*\d+", f"目标OVR {target_overall}", base["note"])
        note = re.sub(r"近邻：[^；]+$", f"近邻：{'、'.join(neighbors)}", note)

        result = {
            **base,
            **identity,
            "order": order,
            "source": source,
            "attributes": attributes,
            "grade": builder.grade_for(target_overall),
            "pool": builder.role_pool(role),
            "processing": processing,
            "inputCoverage": input_coverage,
            "regressionFields": 26 - min(26, input_coverage),
            "neighbors": neighbors,
            "calculatedOverall": builder.calculated_overall(attributes, role),
            "knownAverageDelta": round(sum(known_deltas) / len(known_deltas), 3) if known_deltas else None,
            "capCount": sum(value == builder.ATTRIBUTE_CAP for value in attributes.values()),
            "note": note,
        }
        if result["calculatedOverall"] != target_overall:
            raise RuntimeError(f"{input_name} OVR 耦合失败")
        results.append(result)

        values[0] = order
        values[14] = target_overall
        values[15] = result["grade"]
        values[16] = result["pool"]
        values[17] = role
        values[18] = secondary_role or None
        for index, key in enumerate(builder.ATTRS):
            values[24 + index] = attributes[key]
        values[53] = note
        player_rows.append(values)

    builder.replace_sheet_rows(player_sheet, player_rows)
    player_sheet.data_validations.dataValidation = []
    last_row = len(results) + 1
    for validation_type, formula, sqref in [
        ("list", '"S,A,B,C,"', f"P2:P{last_row}"),
        ("list", '"GK,DEF,MID,ATT,"', f"Q2:Q{last_row}"),
        ("list", '"GK,CB,LB,RB,DM,AM,LM,RM,ST,LW,RW,"', f"R2:S{last_row}"),
        ("list", '"左脚,右脚"', f"X2:X{last_row}"),
    ]:
        validation = DataValidation(type=validation_type, formula1=formula, allow_blank=True)
        player_sheet.add_data_validation(validation)
        validation.add(sqref)
    numeric = DataValidation(type="whole", operator="between", formula1="1", formula2="96", allow_blank=False)
    player_sheet.add_data_validation(numeric)
    numeric.add(f"O2:O{last_row}")
    numeric.add(f"Y2:AX{last_row}")

    source_rows = [[
        result["order"], result["input"], result["nationality"], result["processing"],
        result["inputCoverage"], result["regressionFields"], result["source"].get("sourceId") or None,
        "、".join(result["neighbors"]), result["note"],
    ] for result in results]
    builder.replace_sheet_rows(workbook.worksheets[2], source_rows)

    peak_rows = [[
        result["nameZh"], result["targetOverall"], result["peakSeason"], result["peakAge"], result["peakClub"],
        "左脚" if result["preferredFoot"] == "left" else "右脚",
        result["source"].get("sourceUrl") or builder.wiki_search(result["nameEn"]),
    ] for result in results]
    builder.replace_sheet_rows(workbook.worksheets[3], peak_rows)

    audit_rows = [[
        result["order"], result["input"], result["role"], result.get("secondaryRole") or None,
        result["targetOverall"], result["calculatedOverall"], result["inputCoverage"], result["regressionFields"],
        result["source"].get("overall"), result["processing"], result["knownAverageDelta"], result["capCount"],
        "、".join(result["neighbors"]),
    ] for result in results]
    audit_sheet = workbook.worksheets[4]
    builder.replace_sheet_rows(audit_sheet, audit_rows)
    training_start = len(results) + 4
    for column, label in enumerate(["位置", "训练样本数", "OVR最低", "OVR最高"], start=1):
        audit_sheet.cell(training_start, column, label)
    for offset, role_name in enumerate(["ST", "CB", "DM", "RW", "GK", "AM", "LB", "RB", "LW", "LM"], start=1):
        role_samples = [sample for sample in samples if sample["role"] == role_name]
        audit_sheet.cell(training_start + offset, 1, role_name)
        audit_sheet.cell(training_start + offset, 2, len(role_samples))
        audit_sheet.cell(training_start + offset, 3, min((sample["overall"] for sample in role_samples), default=None))
        audit_sheet.cell(training_start + offset, 4, max((sample["overall"] for sample in role_samples), default=None))

    summary = workbook.worksheets[0]
    summary["A1"] = "第三批 DLC 球员：S4 耦合回归复核版"
    summary["B3"] = len(results)
    summary["B4"] = "已按人工删除结果移除：伊布拉希莫维奇、德容、库伊特"
    summary["B5"] = sum(result["inputCoverage"] == 0 for result in results)
    summary["B6"] = sum(0 < result["inputCoverage"] < 26 for result in results)
    summary["B7"] = sum(result["inputCoverage"] >= 26 for result in results)
    summary["B8"] = " / ".join(str(sum(result["grade"] == grade for result in results)) for grade in ["S", "A", "B", "C"])
    summary["B10"] = builder.BUILD_DATE
    summary["B15"] = "以用户编辑后的 47 人名单和调整后 OVR 为准重新回归；26 项属性已重新耦合，属性计算 OVR 与目标完全一致；基础属性范围 1–96"

    payload = {
        "schemaVersion": 1,
        "generatedAt": builder.BUILD_DATE,
        "basedOn": INPUT_XLSX.name,
        "officialApi": builder.EA_API,
        "players": results,
    }
    workbook.save(OUTPUT_XLSX)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "xlsx": str(OUTPUT_XLSX),
        "json": str(OUTPUT_JSON),
        "players": len(results),
        "overallChanges": [
            [result["input"], base_by_name[result["input"]]["targetOverall"], result["targetOverall"]]
            for result in results
            if base_by_name[result["input"]]["targetOverall"] != result["targetOverall"]
        ],
        "deleted": [player["input"] for player in base_payload["players"] if player["input"] not in {result["input"] for result in results}],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
